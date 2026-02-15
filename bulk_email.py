import os
import time
from email.message import EmailMessage
from email.utils import formataddr
import smtplib
from dotenv import load_dotenv
from openpyxl import load_workbook
from mako.template import Template


load_dotenv()
smtp_server = "smtp.gmail.com"
smtp_port = 587
login_id = os.environ.get("LOGIN_ID", "")
sender_name = os.environ.get("SENDER_NAME", "")
pwd = os.environ.get("APP_PASSWORD", "")

html_template = Template(filename="meeting_link.html")


def send_smtp(
    smtp_server: str,
    smtp_port: int,
    login_id: str,
    pwd: str,
    sender_name: str,
    recipients_list: list[tuple[str, str, bool]],
    subject: str,
    html_template: Template,
    start: int = 1,
    count: int = -1,
    sleep_sec: int = 1,
    dry_run: bool = True,
):
    msg = EmailMessage()
    msg["From"] = formataddr((sender_name, login_id))
    msg["To"] = formataddr((recipients_list[0][1], recipients_list[0][0]))
    msg["Subject"] = subject

    if start < 1:
        start = 1

    if count < 0 or count > len(recipients_list):
        count = len(recipients_list) - start + 1

    if dry_run:
        print("Dry run: Emails will not be sent. Here are the details:")
        print(f"SMTP Server: {smtp_server}:{smtp_port}")
        print(f"Login ID: {login_id} ({sender_name})")

        if count > 0:
            sent_count = 0
            for i, recipient in enumerate(recipients_list, start=1):
                if i < start:
                    continue
                to_email, to_name, att_mode, *_ = recipient
                sent_count += 1
                print(f"{i:4}: {to_name:30} {'<' + to_email + '>':50} {att_mode}")
                if sent_count == count:
                    break
        print(f"Total emails sent: {count}")
        return  # Return from the function after a dry run

    # Strat a real run
    with smtplib.SMTP(smtp_server, smtp_port) as server:
        print("Sending emails")
        server.ehlo()
        server.starttls()
        server.login(login_id, pwd)

        if count > 0:
            sent_count = 0
            for i, recipient in enumerate(recipients_list, start=1):
                if i < start:
                    continue
                to_email, to_name, att_mode, *_ = recipient
                print(
                    f"Sending to: {i:4}: {to_name:30} {to_email:50} {att_mode}",
                    end=" ",
                )
                html = html_template.render(name=to_name, mode=att_mode)
                msg.replace_header("To", f"{to_name} <{to_email}>")
                msg.set_content(html, subtype="html")

                try:
                    server.send_message(msg)
                    sent_count += 1
                    print("Sent")
                    if sent_count == count:
                        break

                    time.sleep(sleep_sec)
                except Exception as e:
                    print(f"An error occurred while sending the email: {e}")
        print(f"Total emails sent: {i}")


def clean_name(name: str) -> str:
    name = name.strip()
    hon = name.split(".", 1)
    if len(hon) > 1 and hon[0] in ["Prof", "Dr", "Mr", "Ms", "Er"]:
        name = f"{hon[0]}. {hon[1]}"
    name = " ".join(w.capitalize() if len(w) != 2 else w.upper() for w in name.split())
    return name


def prepare_recipients_list_from_excel(fname: str) -> list[tuple[str, str, bool]]:
    rec_lst = []
    wb = load_workbook(fname)
    ws = wb.active
    if ws:
        for row in ws.iter_rows(min_row=2, values_only=True):
            email = str(row[1]).strip()
            name = clean_name(str(row[2]))
            att_mode = (
                str(row[4]).strip().split("(", 1)[0].strip().replace(" ", "").lower()
            )
            rec_lst.append(tuple((email, name, att_mode == "online")))
    return rec_lst


if __name__ == "__main__":
    subject = "Thank you for registering for the SEA Tech Talk February 2026"

    recipients_list = [
        ("satish.annigeri@outlook.com", "Satish Annigeri", True),
        ("satish.annigeri@outlook.com", "Satish Annigeri", False),
        # ("asifdanwad@gmail.com", "Asif Iqbal Danwad", True),
        # ("asifdanwad@gmail.com", "Asif Iqbal Danwad", False),
    ]

    rec_lst = prepare_recipients_list_from_excel("python_for_str_engg.xlsx")
    recipients_list = recipients_list + rec_lst
    for rec in recipients_list:
        print(rec)

    send_smtp(
        smtp_server,
        smtp_port,
        login_id,
        pwd,
        sender_name,
        recipients_list,
        subject,
        html_template,
        start=1,
        count=1,
        dry_run=True,
    )
